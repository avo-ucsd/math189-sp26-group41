"""
Label every game in the free / paid lists by its listed Steam release date and
flag whether it falls inside the project's 2011-2018 study window.

Output: eda/release_window_audit.xlsx
  - Free Games      : one row per free-to-play game, sorted by release date
  - Paid Games      : one row per paid game, sorted by release date
  - Summary         : in/out counts per group + year distribution (Excel formulas)

The "Release Date" column is taken verbatim from the scraped list files. For some
re-released titles (e.g. "Grand Theft Auto V Legacy") this store date differs from
the game's original launch; such rows are flagged in the Note column.

Run from the repo root:  python eda/label_release_window.py
"""
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

REPO = Path(__file__).resolve().parents[1]
DATA = REPO / "data"
OUT = REPO / "eda" / "release_window_audit.xlsx"

WIN_LO, WIN_HI = 2011, 2018
# substrings that indicate a re-release / repackaged store entry whose listed
# date may not equal the game's original launch date
EDITION_HINTS = ["Legacy", "Complete", "Pack", "Aeternum", "Master Chief Collection"]

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
OUT_FILL = PatternFill("solid", fgColor="F8CBAD")   # light red - outside window
IN_FILL = PatternFill("solid", fgColor="C6E0B4")    # light green - inside window
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")  # light yellow - caveat
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load(group):
    f = "top50_free_games_list.csv" if group == "free" else "top50_paid_games_list.csv"
    df = pd.read_csv(DATA / f)
    df["Release Date"] = pd.to_datetime(df["Release Date"], errors="coerce")
    if "Base Price" not in df.columns:
        df["Base Price"] = "Free"
    df["Note"] = df["Game Title"].apply(
        lambda t: "Re-release/edition date - verify original launch"
        if any(h.lower() in str(t).lower() for h in EDITION_HINTS) else ""
    )
    return df.sort_values("Release Date").reset_index(drop=True)


def write_games_sheet(wb, title, df):
    ws = wb.create_sheet(title)
    cols = ["AppID", "Game Title", "Release Date", "Release Year",
            "In Window (2011-2018)", "Base Price", "Estimated Owners",
            "Peak Concurrent Players (24h)", "Note"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF")
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    for i, r in df.iterrows():
        rx = i + 2
        ws.cell(rx, 1, int(r["AppID"]))
        ws.cell(rx, 2, r["Game Title"])
        dc = ws.cell(rx, 3, r["Release Date"].to_pydatetime() if pd.notna(r["Release Date"]) else None)
        dc.number_format = "yyyy-mm-dd"
        ws.cell(rx, 4, f"=YEAR(C{rx})")
        ws.cell(rx, 5, f'=IF(AND(D{rx}>={WIN_LO},D{rx}<={WIN_HI}),"In","Out")')
        ws.cell(rx, 6, r["Base Price"])
        ws.cell(rx, 7, r["Estimated Owners"])
        ws.cell(rx, 8, r["Peak Concurrent Players (24h)"])
        ncell = ws.cell(rx, 9, r["Note"])
        yr = r["Release Date"].year if pd.notna(r["Release Date"]) else None
        in_win = yr is not None and WIN_LO <= yr <= WIN_HI
        for c in range(1, len(cols) + 1):
            cell = ws.cell(rx, c)
            cell.font = Font(name=FONT)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center" if c != 2 else "left")
        ws.cell(rx, 5).fill = IN_FILL if in_win else OUT_FILL
        if r["Note"]:
            ncell.fill = NOTE_FILL

    widths = [10, 34, 14, 12, 18, 12, 18, 16, 40]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = w
    ws.freeze_panes = "A2"
    return ws, len(df)


def write_summary(wb, n_free, n_paid):
    ws = wb.create_sheet("Summary", 0)
    ws["A1"] = "Release-Window Audit: 2011-2018 study scope"
    ws["A1"].font = Font(name=FONT, bold=True, size=14)
    ws["A3"] = ("Each game is labeled by the Release Date listed in the scraped list files. "
                "'In' = released 2011-2018 (project scope); 'Out' = outside scope.")
    ws["A3"].font = Font(name=FONT, italic=True)

    hdr = ["Group", "Total games", "In 2011-2018", "Outside window", "% in window"]
    ws.append([])
    ws.append(hdr)
    hrow = ws.max_row
    for c in range(1, 6):
        cell = ws.cell(hrow, c)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF")
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    fr, pr = hrow + 1, hrow + 2
    free_rng = f"'Free Games'!E2:E{n_free+1}"
    paid_rng = f"'Paid Games'!E2:E{n_paid+1}"
    ws.cell(fr, 1, "Free-to-play")
    ws.cell(fr, 2, n_free)
    ws.cell(fr, 3, f'=COUNTIF({free_rng},"In")')
    ws.cell(fr, 4, f'=COUNTIF({free_rng},"Out")')
    ws.cell(fr, 5, f"=C{fr}/B{fr}")
    ws.cell(pr, 1, "Paid")
    ws.cell(pr, 2, n_paid)
    ws.cell(pr, 3, f'=COUNTIF({paid_rng},"In")')
    ws.cell(pr, 4, f'=COUNTIF({paid_rng},"Out")')
    ws.cell(pr, 5, f"=C{pr}/B{pr}")
    tr = pr + 1
    ws.cell(tr, 1, "TOTAL")
    ws.cell(tr, 2, f"=B{fr}+B{pr}")
    ws.cell(tr, 3, f"=C{fr}+C{pr}")
    ws.cell(tr, 4, f"=D{fr}+D{pr}")
    ws.cell(tr, 5, f"=C{tr}/B{tr}")
    for rr in (fr, pr, tr):
        for c in range(1, 6):
            cell = ws.cell(rr, c)
            cell.font = Font(name=FONT, bold=(rr == tr))
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center" if c > 1 else "left")
        ws.cell(rr, 5).number_format = "0.0%"

    ws.column_dimensions["A"].width = 16
    for c in "BCDE":
        ws.column_dimensions[c].width = 16
    return ws


def main():
    free = load("free")
    paid = load("paid")
    wb = Workbook()
    wb.remove(wb.active)
    _, nf = write_games_sheet(wb, "Free Games", free)
    _, np_ = write_games_sheet(wb, "Paid Games", paid)
    write_summary(wb, nf, np_)
    OUT.parent.mkdir(exist_ok=True)
    wb.save(OUT)
    # console evidence
    for name, df in [("FREE", free), ("PAID", paid)]:
        yr = df["Release Date"].dt.year
        inw = yr.between(WIN_LO, WIN_HI)
        print(f"{name}: total {len(df)} | In {inw.sum()} | Out {(~inw).sum()}")
    print(f"Saved -> {OUT}")


if __name__ == "__main__":
    main()
